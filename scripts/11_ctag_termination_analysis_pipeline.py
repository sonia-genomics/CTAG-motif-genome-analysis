import pandas as pd
import numpy as np
from openpyxl.styles import Alignment, Font
import matplotlib.pyplot as plt
import os

input_tri = "output/final_tri_analysis.xlsx"
input_tetra = "output/final_tetra_analysis.xlsx"
output_file = "output/final_analysis.xlsx"

tri_sheets = pd.read_excel(input_tri, sheet_name=None)
tetra_sheets = pd.read_excel(input_tetra, sheet_name=None)

# ---------- FUNCTIONS ----------
def safe_div(a, b):
    return a / b if b != 0 else np.nan

def get_val(df, motif, col):
    sub = df[df.iloc[:, 0] == motif]
    return sub[col].values[0] if not sub.empty else 0


# =========================================================
# BUILD TABLES
# =========================================================
t4, t5, t6, t7, t8 = [], [], [], [], []

for org in tri_sheets:

    df_tri = tri_sheets[org].copy()
    df_tri.rename(columns={df_tri.columns[0]: "TRINUC"}, inplace=True)

    df_tetra = tetra_sheets.get(org)
    if df_tetra is None:
        continue

    df_tetra.rename(columns={df_tetra.columns[0]: "Tetranucleotide"}, inplace=True)

    # ---------- TABLE 4 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        for c in ["TAG","TAA","TGA"]:
            row.append(get_val(df_tri, c, col))
    t4.append(row)

    # ---------- TABLE 5 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        for c in ["CTAG","CTAA","CTGA"]:
            row.append(get_val(df_tetra, c, col))
    t5.append(row)

    # ---------- TABLE 6 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        r1 = safe_div(get_val(df_tetra,"CTAA",col)+get_val(df_tetra,"CTGA",col),
                      get_val(df_tetra,"CTAG",col))
        r2 = safe_div(get_val(df_tri,"TAA",col)+get_val(df_tri,"TGA",col),
                      get_val(df_tri,"TAG",col))
        final = safe_div(r1, r2)
        row += [r1, r2, final]
    t6.append(row)

    # ---------- TABLE 7 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        for c in ["CTAG","GTAG","ATAG","TTAG"]:
            row.append(get_val(df_tetra, c, col))
    t7.append(row)

    # ---------- TABLE 8 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        r1 = safe_div(get_val(df_tetra,"CTAA",col)+get_val(df_tetra,"CTGA",col),
                      get_val(df_tetra,"CTAG",col))
        r2 = safe_div(get_val(df_tetra,"GTAA",col)+get_val(df_tetra,"GTGA",col),
                      get_val(df_tetra,"GTAG",col))
        final = safe_div(r1, r2)
        row += [r1, r2, final]
    t8.append(row)


# Convert to DataFrames
t4 = pd.DataFrame(t4)
t5 = pd.DataFrame(t5)
t6 = pd.DataFrame(t6)
t7 = pd.DataFrame(t7)
t8 = pd.DataFrame(t8)


# =========================================================
# WRITE TO EXCEL
# =========================================================
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

    t4.to_excel(writer, sheet_name="Table 4", index=False)
    t5.to_excel(writer, sheet_name="Table 5", index=False)
    t6.to_excel(writer, sheet_name="Table 6", index=False)
    t7.to_excel(writer, sheet_name="Table 7", index=False)
    t8.to_excel(writer, sheet_name="Table 8", index=False)

    wb = writer.book

    # =====================================================
    # BASIC TABLE FORMAT (4,5,7)
    # =====================================================
    def format_basic(ws, title, codons):

        ws.insert_rows(1, 3)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1+len(codons)*3)
        ws.cell(1,1).value = title
        ws.cell(1,1).font = Font(bold=True)
        ws.cell(1,1).alignment = Alignment(horizontal='center')

        regions = ["GENOME","CODING SEQUENCE","TERMINATION SITE"]

        col = 2
        for r in regions:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+len(codons)-1)
            ws.cell(2,col).value = r
            ws.cell(2,col).alignment = Alignment(horizontal='center')
            col += len(codons)

        col = 2
        for _ in regions:
            for c in codons:
                ws.cell(3,col).value = c
                ws.cell(3,col).alignment = Alignment(horizontal='center')
                col += 1



    # TABLE 6 FORMAT
    def format_table6(ws):

        ws.insert_rows(1, 5)

        ws.merge_cells('A1:J1')
        ws['A1'] = "Ratio: (CTAA+CTGA)/CTAG vs (TAA+TGA)/TAG"
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(bold=True)

        ws.merge_cells('B3:D3')
        ws['B3'] = "O/E (CTAA + CTGA) / O/E (CTAG)"

        ws.merge_cells('E3:G3')
        ws['E3'] = "O/E (TAA + TGA) / O/E (TAG)"

        ws.merge_cells('H3:J3')
        ws['H3'] = "Final Ratio"

        regions = ["GENOME","CODING REGION","TERMINATION SITE"]

        col = 2
        for _ in range(3):
            for i, r in enumerate(regions):
                ws.cell(4, col+i).value = r
                ws.cell(4, col+i).alignment = Alignment(horizontal='center')
            col += 3


    # TABLE 8 FORMAT
    def format_table8(ws):

        ws.insert_rows(1, 5)

        ws.merge_cells('A1:J1')
        ws['A1'] = "Ratio: CTAG vs DTAG"
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(bold=True)

        ws.merge_cells('B3:D3')
        ws['B3'] = "O/E (CTAA + CTGA) / O/E (CTAG)"

        ws.merge_cells('E3:G3')
        ws['E3'] = "O/E (DTAA + DTGA) / O/E (DTAG)"

        ws.merge_cells('H3:J3')
        ws['H3'] = "Final Ratio"

        regions = ["GENOME","CODING REGION","TERMINATION SITE"]

        col = 2
        for _ in range(3):
            for i, r in enumerate(regions):
                ws.cell(4, col+i).value = r
                ws.cell(4, col+i).alignment = Alignment(horizontal='center')
            col += 3


    # APPLY FORMATTING
    format_basic(wb["Table 4"], "TERMINATION CODON", ["TAG","TAA","TGA"])
    format_basic(wb["Table 5"], "CYTOSINE PRECEDING TERMINATION CODON", ["CTAG","CTAA","CTGA"])
    format_basic(wb["Table 7"], "O/E VALUE OF DTAG", ["CTAG","GTAG","ATAG","TTAG"])

    format_table6(wb["Table 6"])
    format_table8(wb["Table 8"])


print("\nSUCCESS: ALL TABLES GENERATED!")
print("Output:", output_file)



#------------------------------------------------------------------


plot_dir = os.path.join(os.path.dirname(output_file), "plots")
os.makedirs(plot_dir, exist_ok=True)

# =========================================================
# FUNCTION TO EXTRACT REGIONS
# =========================================================
def extract_regions(df):
    data = df.iloc[:, 1:]  # remove organism column
    n = data.shape[1] // 3

    genome = data.iloc[:, :n].values.flatten()
    coding = data.iloc[:, n:2*n].values.flatten()
    term = data.iloc[:, 2*n:3*n].values.flatten()

    # Remove NaN
    genome = genome[~np.isnan(genome)]
    coding = coding[~np.isnan(coding)]
    term = term[~np.isnan(term)]

    return genome, coding, term


# =========================================================
# PLOT FUNCTION
# =========================================================
def make_boxplot(df, title, filename):

    g, c, t = extract_regions(df)

    plt.figure()
    plt.boxplot([g, c, t], labels=["GENOME", "CODING REGION", "TERMINATION SITE"])
    plt.title(title)
    plt.xlabel("Regions")
    plt.ylabel("O/E Values / Ratios")

    save_path = os.path.join(plot_dir, filename)
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"Plot saved: {save_path}")


# =========================================================
# GENERATE FOR ALL TABLES
# =========================================================
make_boxplot(t4, "Table 4: Termination Codons", "table4_boxplot.png")
make_boxplot(t5, "Table 5: Cytosine Preference", "table5_boxplot.png")
make_boxplot(t6, "Table 6: CTAG Ratio", "table6_boxplot.png")
make_boxplot(t7, "Table 7: DTAG", "table7_boxplot.png")
make_boxplot(t8, "Table 8: DTAG Ratio", "table8_boxplot.png")