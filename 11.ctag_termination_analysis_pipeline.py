import pandas as pd
import numpy as np
from openpyxl.styles import Alignment, Font

input_tri = "output/final_tri_analysis.xlsx"
input_tetra = "output/final_CTAg_analysis.xlsx"
output_file = "output/final_analysis.xlsx"


tri_sheets = pd.read_excel(input_tri, sheet_name=None)
tetra_sheets = pd.read_excel(input_tetra, sheet_name=None)


def safe_div(a, b):
    return a / b if b != 0 else np.nan

def get_val(df, motif, col):
    sub = df[df.iloc[:, 0] == motif]
    return sub[col].values[0] if not sub.empty else 0


t4, t5, t6, t7, t8 = [], [], [], [], []

for org in tri_sheets:

    df_tri = tri_sheets[org].copy()
    df_tri.rename(columns={df_tri.columns[0]: "TRINUC"}, inplace=True)

    df_tetra = tetra_sheets.get(org)
    if df_tetra is None:
        continue

    df_tetra.rename(columns={df_tetra.columns[0]: "Tetranucleotide"}, inplace=True)

    # ---------- TABLE 4 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        for c in ["TAG","TAA","TGA"]:
            row.append(get_val(df_tri, c, col))
    t4.append(row)

    # ---------- TABLE 5 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        for c in ["CTAG","CTAA","CTGA"]:
            row.append(get_val(df_tetra, c, col))
    t5.append(row)

    # ---------- TABLE 6 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        r1 = safe_div(get_val(df_tetra,"CTAA",col)+get_val(df_tetra,"CTGA",col),
                      get_val(df_tetra,"CTAG",col))
        r2 = safe_div(get_val(df_tri,"TAA",col)+get_val(df_tri,"TGA",col),
                      get_val(df_tri,"TAG",col))
        final = safe_div(r1, r2)
        row += [r1, r2, final]
    t6.append(row)

    # ---------- TABLE 7 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        for c in ["CTAG","GTAG","ATAG","TTAG"]:
            row.append(get_val(df_tetra, c, col))
    t7.append(row)

    # ---------- TABLE 8 ----------
    row = [org]
    for col in ["Genome_OE","Coding_OE","Term_OE"]:
        r1 = safe_div(get_val(df_tetra,"CTAA",col)+get_val(df_tetra,"CTGA",col),
                      get_val(df_tetra,"CTAG",col))
        r2 = safe_div(get_val(df_tetra,"GTAA",col)+get_val(df_tetra,"GTGA",col),
                      get_val(df_tetra,"GTAG",col))
        final = safe_div(r1, r2)
        row += [r1, r2, final]
    t8.append(row)


# Convert to DataFrames
t4 = pd.DataFrame(t4)
t5 = pd.DataFrame(t5)
t6 = pd.DataFrame(t6)
t7 = pd.DataFrame(t7)
t8 = pd.DataFrame(t8)

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

    t4.to_excel(writer, sheet_name="Table 4", index=False)
    t5.to_excel(writer, sheet_name="Table 5", index=False)
    t6.to_excel(writer, sheet_name="Table 6", index=False)
    t7.to_excel(writer, sheet_name="Table 7", index=False)
    t8.to_excel(writer, sheet_name="Table 8", index=False)

    wb = writer.book


    def format_basic(ws, title, codons):

        ws.insert_rows(1, 3)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1+len(codons)*3)
        ws.cell(1,1).value = title
        ws.cell(1,1).font = Font(bold=True)
        ws.cell(1,1).alignment = Alignment(horizontal='center')

        regions = ["GENOME","CODING SEQUENCE","TERMINATION SITE"]

        col = 2
        for r in regions:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+len(codons)-1)
            ws.cell(2,col).value = r
            ws.cell(2,col).alignment = Alignment(horizontal='center')
            col += len(codons)

        col = 2
        for _ in regions:
            for c in codons:
                ws.cell(3,col).value = c
                ws.cell(3,col).alignment = Alignment(horizontal='center')
                col += 1


    def format_table6(ws):

        ws.insert_rows(1, 5)

        ws.merge_cells('A1:J1')
        ws['A1'] = "Ratio: (CTAA+CTGA)/CTAG vs (TAA+TGA)/TAG"
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(bold=True)

        ws.merge_cells('B3:D3')
        ws['B3'] = "O/E (CTAA + CTGA) / O/E (CTAG)"

        ws.merge_cells('E3:G3')
        ws['E3'] = "O/E (TAA + TGA) / O/E (TAG)"

        ws.merge_cells('H3:J3')
        ws['H3'] = "Final Ratio"

        regions = ["GENOME","CODING REGION","TERMINATION SITE"]

        col = 2
        for _ in range(3):
            for i, r in enumerate(regions):
                ws.cell(4, col+i).value = r
                ws.cell(4, col+i).alignment = Alignment(horizontal='center')
            col += 3


    def format_table8(ws):

        ws.insert_rows(1, 5)

        ws.merge_cells('A1:J1')
        ws['A1'] = "Ratio: CTAG vs DTAG"
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(bold=True)

        ws.merge_cells('B3:D3')
        ws['B3'] = "O/E (CTAA + CTGA) / O/E (CTAG)"

        ws.merge_cells('E3:G3')
        ws['E3'] = "O/E (DTAA + DTGA) / O/E (DTAG)"

        ws.merge_cells('H3:J3')
        ws['H3'] = "Final Ratio"

        regions = ["GENOME","CODING REGION","TERMINATION SITE"]

        col = 2
        for _ in range(3):
            for i, r in enumerate(regions):
                ws.cell(4, col+i).value = r
                ws.cell(4, col+i).alignment = Alignment(horizontal='center')
            col += 3


    # APPLY FORMATTING
    format_basic(wb["Table 4"], "TERMINATION CODON", ["TAG","TAA","TGA"])
    format_basic(wb["Table 5"], "CYTOSINE PRECEDING TERMINATION CODON", ["CTAG","CTAA","CTGA"])
    format_basic(wb["Table 7"], "O/E VALUE OF DTAG", ["CTAG","GTAG","ATAG","TTAG"])

    format_table6(wb["Table 6"])
    format_table8(wb["Table 8"])


print("\n SUCCESS: ALL TABLES GENERATED!")
print("Output:", output_file)


